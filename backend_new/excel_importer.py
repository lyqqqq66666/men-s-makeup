import os
from typing import Dict, List

import openpyxl

import db_manager


EXCEL_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'Beauty_Product_Data_Structure.xlsx')


def _split_colors(value) -> List[str]:
    text = str(value or '').strip()
    if not text:
        return []
    return [item.strip() for item in text.split(',') if item.strip()]


def _sheet_rows(sheet):
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell or '').strip() for cell in rows[0]]
    result = []
    for row in rows[1:]:
        item = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            item[header] = row[idx] if idx < len(row) else None
        if any(v not in (None, '') for v in item.values()):
            result.append(item)
    return result


def _load_workbook(path=EXCEL_PATH):
    if not os.path.exists(path):
        raise FileNotFoundError(f'Excel文件不存在: {path}')
    return openpyxl.load_workbook(path, data_only=True)


def parse_excel(path=EXCEL_PATH) -> Dict[str, List[Dict]]:
    wb = _load_workbook(path)

    season_rule_rows = _sheet_rows(wb['色彩规则表'])
    sku_rows = _sheet_rows(wb['SKU Database'])
    season_sku_rows = _sheet_rows(wb['季型 - SKU 映射表'])
    recommended_palette_rows = _sheet_rows(wb['推荐色系映射表'])

    recommended_palette_map = {
        str(row.get('season_type') or '').strip(): _split_colors(row.get('palette_colors'))
        for row in recommended_palette_rows
        if str(row.get('season_type') or '').strip()
    }

    season_rules = []
    for row in season_rule_rows:
        season_type = str(row.get('season_type') or '').strip()
        if not season_type:
            continue
        season_rules.append({
            'season_type': season_type,
            'tone': str(row.get('tone') or '').strip() or None,
            'lips_palette': _split_colors(row.get('lips_palette')),
            'cheeks_palette': _split_colors(row.get('cheeks_palette')),
            'brow_palette': _split_colors(row.get('brow_palette')),
            'avoid_palette': _split_colors(row.get('avoid_palette')),
            'recommended_palette': recommended_palette_map.get(season_type, []),
        })

    area_to_category = {
        'lips': 'lip',
        'cheeks': 'contour',
        'brow': 'brow',
        'skin': 'base',
        'eye': 'eye',
    }

    products = []
    for row in sku_rows:
        sku_id = str(row.get('sku_id') or '').strip()
        if not sku_id:
            continue
        apply_area = str(row.get('apply_area') or '').strip()
        category = area_to_category.get(apply_area, 'lip')
        shade_name = str(row.get('shade_name') or '').strip()
        hex_color = str(row.get('hex_color') or '').strip().upper()
        products.append({
            'p_id': sku_id,
            'product_group_id': str(row.get('product_name') or sku_id).strip().lower().replace(' ', '_'),
            'name': str(row.get('product_name') or sku_id).strip(),
            'brand': str(row.get('brand') or '').strip() or 'UNKNOWN',
            'category': category,
            'apply_area': apply_area or 'lips',
            'render_hex': hex_color,
            'render_mode': 1,
            'finish_type': 'semi-matte',
            'opacity': float(row.get('opacity') or 0.6),
            'feather': max(1, int(round(float(row.get('feather') or 0.15) * 100))),
            'transparency_max': float(row.get('transparency_max') or 0.7),
            'season_tag': str(row.get('season_match') or '').strip() or None,
            'color_hex': hex_color,
            'price': float(row.get('price') or 0),
            'image_url': None,
            'mask_params': {},
            'render_params': {'source': 'excel', 'shade_name': shade_name},
            'stock': 100,
            'color_profiles': [
                {
                    'shade_name': shade_name,
                    'hex_color': hex_color,
                    'color_role': 'primary',
                }
            ]
        })

    season_sku_map = []
    for row in season_sku_rows:
        season_type = str(row.get('season_type') or '').strip()
        product_id = str(row.get('recommended_sku') or '').strip()
        if season_type and product_id:
            season_sku_map.append({'season_type': season_type, 'product_id': product_id})

    return {
        'season_rules': season_rules,
        'products': products,
        'season_sku_map': season_sku_map,
    }


def import_excel_to_db(path=EXCEL_PATH):
    payload = parse_excel(path)
    job_id = db_manager.create_import_job(source_name='Beauty_Product_Data_Structure.xlsx', source_path=path, status='running')
    try:
        db_manager.replace_excel_products(payload['products'])
        db_manager.replace_season_rules(payload['season_rules'])
        db_manager.replace_season_sku_map(payload['season_sku_map'])
        summary = {
            'products': len(payload['products']),
            'season_rules': len(payload['season_rules']),
            'season_sku_map': len(payload['season_sku_map']),
        }
        db_manager.finish_import_job(job_id, status='success', summary=summary)
        return {'job_id': job_id, **summary}
    except Exception as exc:
        db_manager.finish_import_job(job_id, status='failed', summary=payload, error_message=str(exc))
        raise


if __name__ == '__main__':
    result = import_excel_to_db()
    print(result)
